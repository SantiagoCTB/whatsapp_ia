from flask import Blueprint, render_template, request, redirect, session, url_for, jsonify
from services.db import get_connection
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from config import Config
import os
import uuid
import re
import requests
import json

config_bp = Blueprint('configuracion', __name__)
MEDIA_ROOT = Config.MEDIA_ROOT
os.makedirs(Config.MEDIA_ROOT, exist_ok=True)

# El comodín '*' en `input_text` permite avanzar al siguiente paso sin validar
# la respuesta del usuario. Si es la única regla de un paso se ejecuta
# automáticamente; si coexiste con otras, actúa como respuesta por defecto.

def _require_admin():
    # Debe haber usuario logueado y el rol 'admin' en la lista de roles
    return "user" in session and 'admin' in (session.get('roles') or [])


def _normalize_input(text):
    """Normaliza una lista separada por comas."""
    return ','.join(t.strip().lower() for t in (text or '').split(',') if t.strip())

def _url_ok(url):
    try:
        r = requests.head(url, allow_redirects=True, timeout=5)
        ok = r.status_code == 200
        mime = r.headers.get('Content-Type', '').split(';', 1)[0] if ok else None
        return ok, mime
    except requests.RequestException:
        return False, None

def _reglas_view(template_name):
    """Renderiza las vistas de reglas.
    El comodín '*' en `input_text` avanza al siguiente paso sin validar
    la respuesta del usuario; si existen otras reglas, actúa como opción
    por defecto."""
    if not _require_admin():
        return redirect(url_for("auth.login"))

    conn = get_connection()
    c = conn.cursor()
    try:
        # --- Migraciones defensivas de nuevas columnas ---
        c.execute("SHOW COLUMNS FROM reglas LIKE 'rol_keyword';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN rol_keyword VARCHAR(20) NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'calculo';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN calculo TEXT NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'handler';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN handler VARCHAR(50) NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'media_url';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN media_url TEXT NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'media_tipo';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN media_tipo VARCHAR(20) NULL;")
            conn.commit()

        if request.method == 'POST':
            # Importar desde Excel
            if 'archivo' in request.files and request.files['archivo']:
                archivo = request.files['archivo']
                wb = load_workbook(archivo)
                hoja = wb.active
                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    if not fila:
                        continue
                    # Permitir archivos con columnas opcionales
                    datos = list(fila) + [None] * 11
                    step, input_text, respuesta, siguiente_step, tipo, media_url, media_tipo, opciones, rol_keyword, calculo, handler = datos[:11]
                    url_ok = False
                    detected_type = None
                    if media_url:
                        url_ok, detected_type = _url_ok(str(media_url))
                        if not url_ok:
                            media_url = None
                            media_tipo = None
                        else:
                            media_tipo = media_tipo or detected_type
                    if media_tipo:
                        media_tipo = str(media_tipo).split(';', 1)[0]
                    # Normalizar campos clave
                    step = (step or '').strip().lower()
                    input_text = _normalize_input(input_text)
                    siguiente_step = _normalize_input(siguiente_step) or None

                    c.execute(
                        "SELECT id FROM reglas WHERE step = %s AND input_text = %s",
                        (step, input_text)
                    )
                    existente = c.fetchone()
                    if existente:
                        regla_id = existente[0]
                        c.execute(
                            """
                            UPDATE reglas
                               SET respuesta = %s,
                                   siguiente_step = %s,
                                   tipo = %s,
                                   media_url = %s,
                                   media_tipo = %s,
                                   opciones = %s,
                                   rol_keyword = %s,
                                   calculo = %s,
                                   handler = %s
                             WHERE id = %s
                            """,
                            (respuesta, siguiente_step, tipo, media_url, media_tipo, opciones, rol_keyword, calculo, handler, regla_id)
                        )
                        c.execute("DELETE FROM regla_medias WHERE regla_id=%s", (regla_id,))
                        if media_url and url_ok:
                            c.execute(
                                "INSERT INTO regla_medias (regla_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                                (regla_id, media_url, media_tipo),
                            )
                    else:
                        c.execute(
                            "INSERT INTO reglas (step, input_text, respuesta, siguiente_step, tipo, media_url, media_tipo, opciones, rol_keyword, calculo, handler) "
                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                            (step, input_text, respuesta, siguiente_step, tipo, media_url, media_tipo, opciones, rol_keyword, calculo, handler)
                        )
                        regla_id = c.lastrowid
                        if media_url and url_ok:
                            c.execute(
                                "INSERT INTO regla_medias (regla_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                                (regla_id, media_url, media_tipo),
                            )
                conn.commit()
            else:
                # Entrada manual desde formulario
                step = (request.form['step'] or '').strip().lower() or None
                input_text = _normalize_input(request.form['input_text']) or None
                respuesta = request.form['respuesta']
                siguiente_step = _normalize_input(request.form.get('siguiente_step')) or None
                tipo = request.form.get('tipo', 'texto')
                media_files = request.files.getlist('media') or request.files.getlist('media[]')
                media_url_field = request.form.get('media_url')
                medias = []
                for media_file in media_files:
                    if media_file and media_file.filename:
                        filename = secure_filename(media_file.filename)
                        unique = f"{uuid.uuid4().hex}_{filename}"
                        path = os.path.join(MEDIA_ROOT, unique)
                        media_file.save(path)
                        url = url_for('static', filename=f'uploads/{unique}', _external=True)
                        medias.append((url, media_file.mimetype.split(';', 1)[0]))
                if media_url_field:
                    for url in [u.strip() for u in re.split(r'[\n,]+', media_url_field) if u.strip()]:
                        ok, content_type = _url_ok(url)
                        if not ok:
                            return f"URL no válida: {url}", 400
                        medias.append((url, content_type))
                media_url = medias[0][0] if medias else None
                media_tipo = medias[0][1] if medias else None
                opciones = request.form['opciones']
                list_header = request.form.get('list_header')
                list_footer = request.form.get('list_footer')
                list_button = request.form.get('list_button')
                sections_raw = request.form.get('sections')
                if tipo == 'lista':
                    if not opciones:
                        try:
                            sections = json.loads(sections_raw) if sections_raw else []
                        except Exception:
                            sections = []
                        opts = {
                            'header': list_header,
                            'footer': list_footer,
                            'button': list_button,
                            'sections': sections
                        }
                        opciones = json.dumps(opts)
                rol_keyword = request.form.get('rol_keyword')
                calculo = request.form.get('calculo')
                handler = request.form.get('handler')
                regla_id = request.form.get('regla_id')

                if regla_id:
                    c.execute(
                        """
                        UPDATE reglas
                           SET step = %s,
                               input_text = %s,
                               respuesta = %s,
                               siguiente_step = %s,
                               tipo = %s,
                               media_url = %s,
                               media_tipo = %s,
                               opciones = %s,
                               rol_keyword = %s,
                               calculo = %s,
                               handler = %s
                         WHERE id = %s
                        """,
                        (
                            step,
                            input_text,
                            respuesta,
                            siguiente_step,
                            tipo,
                            media_url,
                            media_tipo,
                            opciones,
                            rol_keyword,
                            calculo,
                            handler,
                            regla_id,
                        ),
                    )
                    c.execute("DELETE FROM regla_medias WHERE regla_id=%s", (regla_id,))
                    for url, tipo_media in medias:
                        c.execute(
                            "INSERT INTO regla_medias (regla_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                            (regla_id, url, tipo_media),
                        )
                else:
                    c.execute(
                        "SELECT id FROM reglas WHERE step = %s AND input_text = %s",
                        (step, input_text)
                    )
                    existente = c.fetchone()
                    if existente:
                        regla_id = existente[0]
                        c.execute(
                            """
                            UPDATE reglas
                               SET respuesta = %s,
                                   siguiente_step = %s,
                                   tipo = %s,
                                   media_url = %s,
                                   media_tipo = %s,
                                   opciones = %s,
                                   rol_keyword = %s,
                                   calculo = %s,
                                   handler = %s
                             WHERE id = %s
                            """,
                            (
                                respuesta,
                                siguiente_step,
                                tipo,
                                media_url,
                                media_tipo,
                                opciones,
                                rol_keyword,
                                calculo,
                                handler,
                                regla_id,
                            ),
                        )
                        c.execute("DELETE FROM regla_medias WHERE regla_id=%s", (regla_id,))
                        for url, tipo_media in medias:
                            c.execute(
                                "INSERT INTO regla_medias (regla_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                                (regla_id, url, tipo_media),
                            )
                    else:
                        c.execute(
                            "INSERT INTO reglas (step, input_text, respuesta, siguiente_step, tipo, media_url, media_tipo, opciones, rol_keyword, calculo, handler) "
                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                            (
                                step,
                                input_text,
                                respuesta,
                                siguiente_step,
                                tipo,
                                media_url,
                                media_tipo,
                                opciones,
                                rol_keyword,
                                calculo,
                                handler,
                            ),
                        )
                        regla_id = c.lastrowid
                        for url, tipo_media in medias:
                            c.execute(
                                "INSERT INTO regla_medias (regla_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                                (regla_id, url, tipo_media),
                            )
                conn.commit()

        # Listar todas las reglas
        c.execute(
            """
            SELECT r.id, r.step, r.input_text, r.respuesta, r.siguiente_step, r.tipo,
                   GROUP_CONCAT(m.media_url SEPARATOR '||') AS media_urls,
                   GROUP_CONCAT(m.media_tipo SEPARATOR '||') AS media_tipos,
                   r.opciones, r.rol_keyword, r.calculo, r.handler
              FROM reglas r
              LEFT JOIN regla_medias m ON r.id = m.regla_id
             GROUP BY r.id
             ORDER BY r.step, r.id
            """
        )
        rows = c.fetchall()
        reglas = []
        for row in rows:
            d = {
                'id': row[0],
                'step': row[1],
                'input_text': row[2],
                'respuesta': row[3],
                'siguiente_step': row[4],
                'tipo': row[5],
                'media_urls': (row[6] or '').split('||') if row[6] else [],
                'media_tipos': (row[7] or '').split('||') if row[7] else [],
                'opciones': row[8] or '',
                'rol_keyword': row[9],
                'calculo': row[10],
                'handler': row[11],
                'header': None,
                'button': None,
                'footer': None,
            }
            if d['opciones']:
                try:
                    opts = json.loads(d['opciones'])
                    if isinstance(opts, dict):
                        d['header'] = opts.get('header')
                        d['button'] = opts.get('button')
                        d['footer'] = opts.get('footer')
                except Exception:
                    pass
            reglas.append(d)
        return render_template(template_name, reglas=reglas)
    finally:
        conn.close()

@config_bp.route('/configuracion', methods=['GET', 'POST'])
def configuracion():
    return _reglas_view('configuracion.html')

@config_bp.route('/reglas', methods=['GET', 'POST'])
def reglas():
    return _reglas_view('reglas.html')

@config_bp.route('/eliminar_regla/<int:regla_id>', methods=['POST'])
def eliminar_regla(regla_id):
    if not _require_admin():
        return redirect(url_for("auth.login"))

    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("DELETE FROM reglas WHERE id = %s", (regla_id,))
        conn.commit()
        return redirect(url_for('configuracion.reglas'))
    finally:
        conn.close()

@config_bp.route('/botones', methods=['GET', 'POST'])
def botones():
    if not _require_admin():
        return redirect(url_for("auth.login"))

    conn = get_connection()
    c = conn.cursor()
    try:
        if request.method == 'POST':
            # Importar botones desde Excel
            if 'archivo' in request.files and request.files['archivo']:
                archivo = request.files['archivo']
                wb = load_workbook(archivo)
                hoja = wb.active
                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    if not fila:
                        continue
                    nombre = fila[0]
                    mensaje = fila[1] if len(fila) > 1 else None
                    tipo = fila[2] if len(fila) > 2 else None
                    media_url = fila[3] if len(fila) > 3 else None
                    medias = []
                    if media_url:
                        urls = [u.strip() for u in re.split(r'[\n,]+', str(media_url)) if u and u.strip()]
                        for url in urls:
                            ok, mime = _url_ok(url)
                            if ok:
                                medias.append((url, mime))
                    if mensaje:
                        c.execute(
                            "INSERT INTO botones (nombre, mensaje, tipo) VALUES (%s, %s, %s)",
                            (nombre, mensaje, tipo)
                        )
                        boton_id = c.lastrowid
                        for url, mime in medias:
                            c.execute(
                                "INSERT INTO boton_medias (boton_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                                (boton_id, url, mime)
                            )
                conn.commit()
            # Agregar botón manual
            elif 'mensaje' in request.form:
                nombre = request.form.get('nombre')
                nuevo_mensaje = request.form['mensaje']
                tipo = request.form.get('tipo')
                media_files = request.files.getlist('media')
                medias = []
                for media_file in media_files:
                    if media_file and media_file.filename:
                        filename = secure_filename(media_file.filename)
                        unique = f"{uuid.uuid4().hex}_{filename}"
                        path = os.path.join(MEDIA_ROOT, unique)
                        media_file.save(path)
                        url = url_for('static', filename=f'uploads/{unique}', _external=True)
                        medias.append((url, media_file.mimetype.split(';', 1)[0]))
                media_url = request.form.get('media_url', '')
                urls = [u.strip() for u in re.split(r'[\n,]+', media_url) if u and u.strip()]
                for url in urls:
                    ok, mime = _url_ok(url)
                    if ok:
                        medias.append((url, mime))
                if nuevo_mensaje:
                    c.execute(
                        "INSERT INTO botones (nombre, mensaje, tipo) VALUES (%s, %s, %s)",
                        (nombre, nuevo_mensaje, tipo)
                    )
                    boton_id = c.lastrowid
                    for url, mime in medias:
                        c.execute(
                            "INSERT INTO boton_medias (boton_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                            (boton_id, url, mime)
                        )
                    conn.commit()

        c.execute(
            """
            SELECT b.id, b.mensaje, b.tipo, b.nombre,
                   GROUP_CONCAT(m.media_url SEPARATOR '||') AS media_urls,
                   GROUP_CONCAT(m.media_tipo SEPARATOR '||') AS media_tipos
              FROM botones b
              LEFT JOIN boton_medias m ON b.id = m.boton_id
             GROUP BY b.id
             ORDER BY b.id
            """
        )
        botones = c.fetchall()
        return render_template('botones.html', botones=botones)
    finally:
        conn.close()

@config_bp.route('/eliminar_boton/<int:boton_id>', methods=['POST'])
def eliminar_boton(boton_id):
    if not _require_admin():
        return redirect(url_for("auth.login"))

    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("DELETE FROM botones WHERE id = %s", (boton_id,))
        conn.commit()
        return redirect(url_for('configuracion.botones'))
    finally:
        conn.close()

@config_bp.route('/get_botones')
def get_botones():
    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute(
            """
            SELECT b.id, b.mensaje, b.tipo, b.nombre,
                   GROUP_CONCAT(m.media_url SEPARATOR '||') AS media_urls,
                   GROUP_CONCAT(m.media_tipo SEPARATOR '||') AS media_tipos
              FROM botones b
              LEFT JOIN boton_medias m ON b.id = m.boton_id
             GROUP BY b.id
             ORDER BY b.id
            """
        )
        rows = c.fetchall()
        return jsonify([
            {
                'id': r[0],
                'mensaje': r[1],
                'tipo': r[2],
                'nombre': r[3],
                'media_urls': r[4].split('||') if r[4] else [],
                'media_tipos': r[5].split('||') if r[5] else []
            }
            for r in rows
        ])
    finally:
        conn.close()
